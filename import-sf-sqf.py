"""
Import SF/SQF posted content from spreadsheet into the database.

Usage:
  python scripts/import-sf-sqf.py <path-to-xlsx> <clientId> [--dry-run]

Examples:
  python scripts/import-sf-sqf.py SF_SQF_Link.xlsx clxyz123abc
  python scripts/import-sf-sqf.py SF_SQF_Link.xlsx clxyz123abc --dry-run

Options:
  --dry-run       Preview what would be imported without writing to DB
  --sheets        Comma-separated sheet names (default: SF,SQF)
                  Options: SF, SQF, SF Leaderboard, SQF Leaderboard
  --sf-times      Two posting times for SF  (default: 08:57,13:57)
  --sqf-times     Two posting times for SQF (default: 09:57,10:57)
  --no-dedup      Allow re-inserting records already in DB
"""

import sys
import os
import re
from datetime import datetime, timezone
from pathlib import Path

# ── Load .env ─────────────────────────────────────────────────────────────────

env_path = Path(__file__).parent.parent / ".env"
if env_path.exists():
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

# ── Args ──────────────────────────────────────────────────────────────────────

args = sys.argv[1:]

def get_arg(prefix):
    for a in args:
        if a.startswith(prefix + "="):
            return a.split("=", 1)[1]
    return None

file_path   = args[0] if len(args) > 0 else None
client_id   = args[1] if len(args) > 1 else None
is_dry_run  = "--dry-run" in args
no_dedup    = "--no-dedup" in args

sheets_arg  = get_arg("--sheets")
sf_times_arg  = get_arg("--sf-times")
sqf_times_arg = get_arg("--sqf-times")

selected_sheets = [s.strip() for s in sheets_arg.split(",")] if sheets_arg else ["SF", "SQF"]
sf_times  = [t.strip() for t in (sf_times_arg  or "08:57,13:57").split(",")]
sqf_times = [t.strip() for t in (sqf_times_arg or "09:57,10:57").split(",")]

if not file_path or not client_id:
    print("Usage: python scripts/import-sf-sqf.py <path-to-xlsx> <clientId> [--dry-run]")
    sys.exit(1)

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    sys.exit(1)

# ── Sheet config ──────────────────────────────────────────────────────────────

SHEET_CONFIG = {
    "SF":             {"link_col": 0, "title_col": 1, "date_col": 2, "platform": "facebook", "times": sf_times,          "deliverable_type": "Short Form"},
    "SQF":            {"link_col": 0, "title_col": 2, "date_col": 3, "platform": "youtube",  "times": sqf_times,         "deliverable_type": "Long Form"},
    "SF Leaderboard": {"link_col": 0, "title_col": 1, "date_col": 2, "platform": "facebook", "times": sf_times[:1],      "deliverable_type": "Short Form"},
    "SQF Leaderboard":{"link_col": 0, "title_col": 2, "date_col": 3, "platform": "youtube",  "times": sqf_times[:1],     "deliverable_type": "Long Form"},
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_cell_url(cell) -> str | None:
    """Extract URL from cell — prefers hyperlink target, falls back to value."""
    # Check hyperlink property first (covers label-only cells like "SF 17")
    if cell.hyperlink:
        target = str(cell.hyperlink.target or "").strip()
        if target.startswith("http"):
            return target
    # Fall back to cell value
    val = str(cell.value or "").strip()
    if val.startswith("http"):
        return val
    return None

def build_date(raw, time_str: str) -> datetime | None:
    """Combine a date cell value with a HH:MM time string, stored as-is (no tz conversion)."""
    if raw is None or raw is False:
        return None
    if isinstance(raw, datetime):
        d = raw
    else:
        try:
            d = datetime.fromisoformat(str(raw))
        except Exception:
            return None
    try:
        hh, mm = map(int, time_str.split(":"))
    except Exception:
        return None
    return datetime(d.year, d.month, d.day, hh, mm, 0, tzinfo=timezone.utc)

# ── DB connection ─────────────────────────────────────────────────────────────

import psycopg2
import psycopg2.extras

db_url = get_arg("--db") or os.environ.get("DATABASE_URL", "")
if not db_url:
    db_url = input("🔗 Paste your DATABASE_URL: ").strip()
if not db_url:
    print("❌ No DATABASE_URL provided.")
    sys.exit(1)

# psycopg2 doesn't accept the "postgresql+..." SQLAlchemy prefix
db_url = re.sub(r"^postgresql\+\w+://", "postgresql://", db_url)
# Also handle prisma-style ?schema= params
db_url = re.sub(r"\?schema=\w+", "", db_url)

conn = psycopg2.connect(db_url)
cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ── Verify client ─────────────────────────────────────────────────────────────

cur.execute('SELECT id, name FROM "Client" WHERE id = %s', (client_id,))
client = cur.fetchone()

if not client:
    print(f"\n❌ Client not found: {client_id}")
    print("\nAvailable clients:")
    cur.execute('SELECT id, name FROM "Client" ORDER BY name')
    for c in cur.fetchall():
        print(f"  {c['id']}  {c['name']}")
    sys.exit(1)

print(f"\n📋 Client: {client['name']} ({client['id']})")
print(f"📁 File:   {os.path.abspath(file_path)}")
print(f"📊 Sheets: {', '.join(selected_sheets)}")
print(f"⏰ SF times:  {' and '.join(sf_times)}")
print(f"⏰ SQF times: {' and '.join(sqf_times)}")
if is_dry_run:
    print("🔍 DRY RUN — nothing will be written\n")

# ── Parse workbook ────────────────────────────────────────────────────────────

import openpyxl

wb = openpyxl.load_workbook(file_path)  # keeps hyperlinks intact

records = []

for sheet_name in selected_sheets:
    cfg = SHEET_CONFIG.get(sheet_name)
    if not cfg:
        print(f"⚠️  Unknown sheet: \"{sheet_name}\" — skipping")
        continue
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  Sheet \"{sheet_name}\" not in workbook — skipping")
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2))  # skip header row 1

    sheet_count   = 0
    skipped_nodate  = 0
    skipped_nolink  = 0

    for row in rows:
        if not row:
            continue

        # Safe cell getter
        def cell(col_idx):
            return row[col_idx] if col_idx < len(row) else None

        link_cell  = cell(cfg["link_col"])
        title_cell = cell(cfg["title_col"])
        date_cell  = cell(cfg["date_col"])

        # Skip if no date
        raw_date = date_cell.value if date_cell is not None else None
        if not raw_date or raw_date is False:
            skipped_nodate += 1
            continue

        # Get URL — from hyperlink or value
        url = get_cell_url(link_cell) if link_cell is not None else None
        if not url:
            skipped_nolink += 1
            continue

        title = str(title_cell.value).strip() if title_cell and title_cell.value else "Untitled"

        for time_str in cfg["times"]:
            posted_at = build_date(raw_date, time_str)
            if not posted_at:
                continue
            records.append({
                "client_id":       client_id,
                "title":           title,
                "platform":        cfg["platform"],
                "url":             url,
                "posted_at":       posted_at,
                "deliverable_type": cfg["deliverable_type"],
            })
            sheet_count += 1

    print(f"  ✅ {sheet_name:<16} {sheet_count} records  "
          f"({skipped_nodate} skipped no-date, {skipped_nolink} skipped no-link)")

print(f"\n📦 Total records parsed: {len(records)}")

if not records:
    print("Nothing to import.")
    conn.close()
    sys.exit(0)

# ── Deduplicate ───────────────────────────────────────────────────────────────

to_insert = records

if not no_dedup:
    cur.execute(
        'SELECT url, "postedAt" FROM "PostedContent" WHERE "clientId" = %s',
        (client_id,)
    )
    existing = {(r["url"], r["postedAt"].replace(tzinfo=timezone.utc)) for r in cur.fetchall()}
    to_insert = [r for r in records if (r["url"], r["posted_at"]) not in existing]
    dupes = len(records) - len(to_insert)
    if dupes:
        print(f"🔁 Skipping {dupes} duplicates already in DB")

print(f"📝 Records to insert: {len(to_insert)}")

# ── Dry run preview ───────────────────────────────────────────────────────────

if is_dry_run:
    print("\n--- DRY RUN PREVIEW (first 10) ---")
    for i, r in enumerate(to_insert[:10], 1):
        ts = r["posted_at"].strftime("%Y-%m-%dT%H:%M")
        print(f"  {i:>2}. [{r['platform']}] {ts}  {r['title'][:70]}")
    print("\n✅ Dry run complete — no data written.")
    conn.close()
    sys.exit(0)

if not to_insert:
    print("✅ Nothing new to import.")
    conn.close()
    sys.exit(0)

# ── Insert ────────────────────────────────────────────────────────────────────

import uuid

CHUNK = 500
imported = 0

for i in range(0, len(to_insert), CHUNK):
    chunk = to_insert[i:i + CHUNK]
    values = [
        (
            str(uuid.uuid4()),
            r["client_id"],
            r["title"],
            r["platform"],
            r["url"],
            r["posted_at"],
            r["deliverable_type"],
        )
        for r in chunk
    ]
    psycopg2.extras.execute_values(
        cur,
        """
        INSERT INTO "PostedContent" (id, "clientId", title, platform, url, "postedAt", "deliverableType", "createdAt")
        VALUES %s
        ON CONFLICT DO NOTHING
        """,
        [(v[0], v[1], v[2], v[3], v[4], v[5], v[6], datetime.now(timezone.utc)) for v in values],
        template="(%s, %s, %s, %s, %s, %s, %s, %s)"
    )
    conn.commit()
    imported += len(chunk)
    print(f"\r⬆️  Inserting... {imported}/{len(to_insert)}", end="", flush=True)

print(f"\n\n✅ Done! Imported {imported} records for \"{client['name']}\".")
conn.close()